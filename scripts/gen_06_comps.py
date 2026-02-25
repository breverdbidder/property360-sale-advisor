"""
Generate 06_valuation_comps.xlsx for Palm Bay Palms Apartments case study.
Three sheets: Comparable Sales, Valuation Scenarios, Buyer Underwriting.
"""
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data import *

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Reusable style objects
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Arial", size=10, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
INPUT_FONT = Font(name="Arial", size=10, color="1565C0")  # Blue for input cells
INPUT_BOLD_FONT = Font(name="Arial", size=10, bold=True, color="1565C0")
SECTION_FONT = Font(name="Arial", size=11, bold=True, color=NAVY)
CURRENCY_FMT = '$#,##0'
PCT_FMT = '0.0%'
PCT_FMT_2 = '0.00%'
GRM_FMT = '0.0'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row, num_cols):
    """Apply navy background, white bold font to a header row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_body_cell(cell, bold=False, input_cell=False):
    """Apply default body font and border."""
    if input_cell:
        cell.font = INPUT_BOLD_FONT if bold else INPUT_FONT
    elif bold:
        cell.font = BOLD_FONT
    else:
        cell.font = BODY_FONT
    cell.border = THIN_BORDER


def style_section_label(cell, text):
    """Apply section header styling."""
    cell.value = text
    cell.font = SECTION_FONT
    cell.border = THIN_BORDER


# ===========================================================================
# Sheet 1: Comparable Sales
# ===========================================================================
def create_comparable_sales_sheet(wb):
    ws = wb.active
    ws.title = "Comparable Sales"

    headers = ["#", "Address", "Units", "Sale Date", "Sale Price",
               "Price/Unit", "Cap Rate", "GRM"]
    col_widths = [5, 35, 8, 12, 15, 14, 12, 8]

    # Write headers in row 1
    for c, header in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=header)
    style_header_row(ws, 1, len(headers))

    # Write comp data (rows 2-6)
    for i, comp in enumerate(COMPS):
        row = i + 2
        num, address, units, date, price, ppu, cap_rate, grm = comp

        ws.cell(row=row, column=1, value=num)
        ws.cell(row=row, column=2, value=address)
        ws.cell(row=row, column=3, value=units)
        ws.cell(row=row, column=4, value=date)
        ws.cell(row=row, column=5, value=price)
        ws.cell(row=row, column=6, value=ppu)
        ws.cell(row=row, column=7, value=cap_rate)
        ws.cell(row=row, column=8, value=grm)

        # Style all cells in this row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c)
            style_body_cell(cell)

        # Number formats
        ws.cell(row=row, column=5).number_format = CURRENCY_FMT
        ws.cell(row=row, column=6).number_format = CURRENCY_FMT
        ws.cell(row=row, column=7).number_format = PCT_FMT
        ws.cell(row=row, column=8).number_format = GRM_FMT
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

    # Averages row (row 7)
    avg_row = len(COMPS) + 2
    ws.cell(row=avg_row, column=1, value="")
    ws.cell(row=avg_row, column=2, value="Averages")

    # Units average
    ws.cell(row=avg_row, column=3, value=f"=AVERAGE(C2:C{avg_row - 1})")
    ws.cell(row=avg_row, column=3).number_format = '0.0'

    # Sale Date - blank
    ws.cell(row=avg_row, column=4, value="")

    # Sale Price average
    ws.cell(row=avg_row, column=5, value=f"=AVERAGE(E2:E{avg_row - 1})")
    ws.cell(row=avg_row, column=5).number_format = CURRENCY_FMT

    # Price/Unit average
    ws.cell(row=avg_row, column=6, value=f"=AVERAGE(F2:F{avg_row - 1})")
    ws.cell(row=avg_row, column=6).number_format = CURRENCY_FMT

    # Cap Rate average
    ws.cell(row=avg_row, column=7, value=f"=AVERAGE(G2:G{avg_row - 1})")
    ws.cell(row=avg_row, column=7).number_format = PCT_FMT

    # GRM average
    ws.cell(row=avg_row, column=8, value=f"=AVERAGE(H2:H{avg_row - 1})")
    ws.cell(row=avg_row, column=8).number_format = GRM_FMT

    # Style averages row as bold
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=avg_row, column=c)
        style_body_cell(cell, bold=True)

    # Column widths
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Freeze top row & auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{avg_row}"

    return ws


# ===========================================================================
# Sheet 2: Valuation Scenarios
# ===========================================================================
def create_valuation_scenarios_sheet(wb):
    ws = wb.create_sheet("Valuation Scenarios")

    col_widths = {"A": 30, "B": 18, "C": 18, "D": 18}
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    # -----------------------------------------------------------------------
    # Section 1: Income Approach (Cap Rate)
    # -----------------------------------------------------------------------
    r = 1
    style_section_label(ws.cell(row=r, column=1), "Income Approach (Cap Rate)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    style_header_row(ws, r, 4)

    r = 2
    # Sub-headers
    ws.cell(row=r, column=1, value="Metric")
    ws.cell(row=r, column=2, value="Low (7.5%)")
    ws.cell(row=r, column=3, value="Mid (7.0%)")
    ws.cell(row=r, column=4, value="High (6.5%)")
    style_header_row(ws, r, 4)

    r = 3
    ws.cell(row=r, column=1, value="NOI (Actual)")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    # Place NOI value in B3 as an input cell
    ws.cell(row=r, column=2, value=NOI_ACTUAL)
    style_body_cell(ws.cell(row=r, column=2), input_cell=True)
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    # Mirror in C3, D3 with formula references
    ws.cell(row=r, column=3, value="=B3")
    style_body_cell(ws.cell(row=r, column=3))
    ws.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws.cell(row=r, column=4, value="=B3")
    style_body_cell(ws.cell(row=r, column=4))
    ws.cell(row=r, column=4).number_format = CURRENCY_FMT

    r = 4
    ws.cell(row=r, column=1, value="Cap Rate")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2, value=0.075)
    style_body_cell(ws.cell(row=r, column=2), input_cell=True)
    ws.cell(row=r, column=2).number_format = PCT_FMT
    ws.cell(row=r, column=3, value=0.070)
    style_body_cell(ws.cell(row=r, column=3), input_cell=True)
    ws.cell(row=r, column=3).number_format = PCT_FMT
    ws.cell(row=r, column=4, value=0.065)
    style_body_cell(ws.cell(row=r, column=4), input_cell=True)
    ws.cell(row=r, column=4).number_format = PCT_FMT

    r = 5
    ws.cell(row=r, column=1, value="Indicated Value")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    # Formulas: Value = NOI / Cap Rate
    ws.cell(row=r, column=2, value="=B3/B4")
    style_body_cell(ws.cell(row=r, column=2), bold=True)
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws.cell(row=r, column=3, value="=C3/C4")
    style_body_cell(ws.cell(row=r, column=3), bold=True)
    ws.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws.cell(row=r, column=4, value="=D3/D4")
    style_body_cell(ws.cell(row=r, column=4), bold=True)
    ws.cell(row=r, column=4).number_format = CURRENCY_FMT

    # -----------------------------------------------------------------------
    # Section 2: GRM Approach (row 7+)
    # -----------------------------------------------------------------------
    r = 7
    style_section_label(ws.cell(row=r, column=1), "GRM Approach (Gross Rent Multiplier)")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    style_header_row(ws, r, 4)

    r = 8
    ws.cell(row=r, column=1, value="Metric")
    ws.cell(row=r, column=2, value="Low (8.0x)")
    ws.cell(row=r, column=3, value="Mid (8.5x)")
    ws.cell(row=r, column=4, value="High (9.0x)")
    style_header_row(ws, r, 4)

    r = 9
    ws.cell(row=r, column=1, value="Gross Annual Rent (at Market)")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2, value=GPR_PROFORMA)
    style_body_cell(ws.cell(row=r, column=2), input_cell=True)
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws.cell(row=r, column=3, value="=B9")
    style_body_cell(ws.cell(row=r, column=3))
    ws.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws.cell(row=r, column=4, value="=B9")
    style_body_cell(ws.cell(row=r, column=4))
    ws.cell(row=r, column=4).number_format = CURRENCY_FMT

    r = 10
    ws.cell(row=r, column=1, value="GRM")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2, value=8.0)
    style_body_cell(ws.cell(row=r, column=2), input_cell=True)
    ws.cell(row=r, column=2).number_format = GRM_FMT
    ws.cell(row=r, column=3, value=8.5)
    style_body_cell(ws.cell(row=r, column=3), input_cell=True)
    ws.cell(row=r, column=3).number_format = GRM_FMT
    ws.cell(row=r, column=4, value=9.0)
    style_body_cell(ws.cell(row=r, column=4), input_cell=True)
    ws.cell(row=r, column=4).number_format = GRM_FMT

    r = 11
    ws.cell(row=r, column=1, value="Indicated Value")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    # Formulas: Value = Gross Rent * GRM
    ws.cell(row=r, column=2, value="=B9*B10")
    style_body_cell(ws.cell(row=r, column=2), bold=True)
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws.cell(row=r, column=3, value="=C9*C10")
    style_body_cell(ws.cell(row=r, column=3), bold=True)
    ws.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws.cell(row=r, column=4, value="=D9*D10")
    style_body_cell(ws.cell(row=r, column=4), bold=True)
    ws.cell(row=r, column=4).number_format = CURRENCY_FMT

    # -----------------------------------------------------------------------
    # Section 3: Price Per Unit Approach (row 13+)
    # -----------------------------------------------------------------------
    r = 13
    style_section_label(ws.cell(row=r, column=1), "Price Per Unit Approach")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    style_header_row(ws, r, 4)

    r = 14
    ws.cell(row=r, column=1, value="Metric")
    ws.cell(row=r, column=2, value="Low ($105K)")
    ws.cell(row=r, column=3, value="Mid ($115K)")
    ws.cell(row=r, column=4, value="High ($125K)")
    style_header_row(ws, r, 4)

    r = 15
    ws.cell(row=r, column=1, value="Total Units")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2, value=TOTAL_UNITS)
    style_body_cell(ws.cell(row=r, column=2), input_cell=True)
    ws.cell(row=r, column=3, value="=B15")
    style_body_cell(ws.cell(row=r, column=3))
    ws.cell(row=r, column=4, value="=B15")
    style_body_cell(ws.cell(row=r, column=4))

    r = 16
    ws.cell(row=r, column=1, value="Price Per Unit")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2, value=105000)
    style_body_cell(ws.cell(row=r, column=2), input_cell=True)
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws.cell(row=r, column=3, value=115000)
    style_body_cell(ws.cell(row=r, column=3), input_cell=True)
    ws.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws.cell(row=r, column=4, value=125000)
    style_body_cell(ws.cell(row=r, column=4), input_cell=True)
    ws.cell(row=r, column=4).number_format = CURRENCY_FMT

    r = 17
    ws.cell(row=r, column=1, value="Indicated Value")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    # Formulas: Value = Units * Price/Unit
    ws.cell(row=r, column=2, value="=B15*B16")
    style_body_cell(ws.cell(row=r, column=2), bold=True)
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT
    ws.cell(row=r, column=3, value="=C15*C16")
    style_body_cell(ws.cell(row=r, column=3), bold=True)
    ws.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws.cell(row=r, column=4, value="=D15*D16")
    style_body_cell(ws.cell(row=r, column=4), bold=True)
    ws.cell(row=r, column=4).number_format = CURRENCY_FMT

    # -----------------------------------------------------------------------
    # Section 4: Recommended List Price (row 19+)
    # -----------------------------------------------------------------------
    r = 19
    style_section_label(ws.cell(row=r, column=1), "Recommended List Price")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    style_header_row(ws, r, 4)

    r = 20
    ws.cell(row=r, column=1, value="Metric")
    ws.cell(row=r, column=2, value="Value")
    style_header_row(ws, r, 4)

    # List Price
    r = 21
    ws.cell(row=r, column=1, value="Recommended List Price")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2, value=ASKING_PRICE)
    style_body_cell(ws.cell(row=r, column=2), bold=True, input_cell=True)
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT

    # $/Unit - formula referencing list price and units
    r = 22
    ws.cell(row=r, column=1, value="Price Per Unit")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2, value="=B21/B15")
    style_body_cell(ws.cell(row=r, column=2), bold=True)
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT

    # Cap Rate on Actual NOI - formula
    r = 23
    ws.cell(row=r, column=1, value="Cap Rate (Actual NOI)")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2, value="=B3/B21")
    style_body_cell(ws.cell(row=r, column=2), bold=True)
    ws.cell(row=r, column=2).number_format = PCT_FMT_2

    # Cap Rate on Pro Forma NOI
    r = 24
    ws.cell(row=r, column=1, value="NOI (Pro Forma)")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2, value=NOI_PROFORMA)
    style_body_cell(ws.cell(row=r, column=2), input_cell=True)
    ws.cell(row=r, column=2).number_format = CURRENCY_FMT

    r = 25
    ws.cell(row=r, column=1, value="Cap Rate (Pro Forma NOI)")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    ws.cell(row=r, column=2, value="=B24/B21")
    style_body_cell(ws.cell(row=r, column=2), bold=True)
    ws.cell(row=r, column=2).number_format = PCT_FMT_2

    # Freeze top area
    ws.freeze_panes = "A2"

    return ws


# ===========================================================================
# Sheet 3: Buyer Underwriting
# ===========================================================================
def create_buyer_underwriting_sheet(wb):
    ws = wb.create_sheet("Buyer Underwriting")

    col_widths = {"A": 28, "B": 18, "C": 18, "D": 18}
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    # -----------------------------------------------------------------------
    # Header row
    # -----------------------------------------------------------------------
    r = 1
    ws.cell(row=r, column=1, value="Buyer Underwriting Analysis")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    style_header_row(ws, r, 4)

    r = 2
    ws.cell(row=r, column=1, value="Metric")
    ws.cell(row=r, column=2, value="Conservative")
    ws.cell(row=r, column=3, value="Base")
    ws.cell(row=r, column=4, value="Aggressive")
    style_header_row(ws, r, 4)

    # -----------------------------------------------------------------------
    # Input Assumptions Section
    # -----------------------------------------------------------------------
    r = 3
    style_section_label(ws.cell(row=r, column=1), "Input Assumptions")
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(2, 5):
        ws.cell(row=r, column=c).border = THIN_BORDER

    # Purchase Price (row 4)
    r = 4
    ws.cell(row=r, column=1, value="Purchase Price")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c in range(2, 5):
        ws.cell(row=r, column=c, value=ASKING_PRICE)
        style_body_cell(ws.cell(row=r, column=c), input_cell=True)
        ws.cell(row=r, column=c).number_format = CURRENCY_FMT

    # Down Payment % (row 5)
    r = 5
    ws.cell(row=r, column=1, value="Down Payment %")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c in range(2, 5):
        ws.cell(row=r, column=c, value=0.25)
        style_body_cell(ws.cell(row=r, column=c), input_cell=True)
        ws.cell(row=r, column=c).number_format = PCT_FMT_2

    # Down Payment $ (row 6) - formula
    r = 6
    ws.cell(row=r, column=1, value="Down Payment $")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c_idx, col_letter in enumerate(["B", "C", "D"], 2):
        ws.cell(row=r, column=c_idx, value=f"={col_letter}4*{col_letter}5")
        style_body_cell(ws.cell(row=r, column=c_idx))
        ws.cell(row=r, column=c_idx).number_format = CURRENCY_FMT

    # Loan Amount (row 7) - formula
    r = 7
    ws.cell(row=r, column=1, value="Loan Amount")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c_idx, col_letter in enumerate(["B", "C", "D"], 2):
        ws.cell(row=r, column=c_idx, value=f"={col_letter}4-{col_letter}6")
        style_body_cell(ws.cell(row=r, column=c_idx))
        ws.cell(row=r, column=c_idx).number_format = CURRENCY_FMT

    # Interest Rate (row 8)
    r = 8
    ws.cell(row=r, column=1, value="Interest Rate")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c in range(2, 5):
        ws.cell(row=r, column=c, value=0.0725)
        style_body_cell(ws.cell(row=r, column=c), input_cell=True)
        ws.cell(row=r, column=c).number_format = PCT_FMT_2

    # Loan Term (row 9)
    r = 9
    ws.cell(row=r, column=1, value="Loan Term (Years)")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c in range(2, 5):
        ws.cell(row=r, column=c, value=30)
        style_body_cell(ws.cell(row=r, column=c), input_cell=True)

    # NOI (row 10)
    r = 10
    ws.cell(row=r, column=1, value="Net Operating Income (NOI)")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    noi_values = [NOI_ACTUAL, 154000, NOI_PROFORMA]
    for c_idx, noi_val in enumerate(noi_values, 2):
        ws.cell(row=r, column=c_idx, value=noi_val)
        style_body_cell(ws.cell(row=r, column=c_idx), input_cell=True)
        ws.cell(row=r, column=c_idx).number_format = CURRENCY_FMT

    # -----------------------------------------------------------------------
    # Output Calculations Section
    # -----------------------------------------------------------------------
    r = 12
    style_section_label(ws.cell(row=r, column=1), "Output Calculations")
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(2, 5):
        ws.cell(row=r, column=c).border = THIN_BORDER

    # Annual Debt Service (row 13) - PMT formula
    # PMT(rate/12, term*12, -loan_amount) * 12
    r = 13
    ws.cell(row=r, column=1, value="Annual Debt Service")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c_idx, col_letter in enumerate(["B", "C", "D"], 2):
        # =PMT(rate/12, term*12, -loan)*12 but openpyxl just writes the formula string
        formula = f"=-PMT({col_letter}8/12,{col_letter}9*12,{col_letter}7)*12"
        ws.cell(row=r, column=c_idx, value=formula)
        style_body_cell(ws.cell(row=r, column=c_idx), bold=True)
        ws.cell(row=r, column=c_idx).number_format = CURRENCY_FMT

    # DSCR (row 14) - NOI / Debt Service
    r = 14
    ws.cell(row=r, column=1, value="Debt Service Coverage Ratio (DSCR)")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c_idx, col_letter in enumerate(["B", "C", "D"], 2):
        ws.cell(row=r, column=c_idx, value=f"={col_letter}10/{col_letter}13")
        style_body_cell(ws.cell(row=r, column=c_idx), bold=True)
        ws.cell(row=r, column=c_idx).number_format = '0.00x'

    # Cash-on-Cash Return (row 15) - (NOI - Debt Service) / Down Payment
    r = 15
    ws.cell(row=r, column=1, value="Cash-on-Cash Return")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c_idx, col_letter in enumerate(["B", "C", "D"], 2):
        ws.cell(row=r, column=c_idx,
                value=f"=({col_letter}10-{col_letter}13)/{col_letter}6")
        style_body_cell(ws.cell(row=r, column=c_idx), bold=True)
        ws.cell(row=r, column=c_idx).number_format = PCT_FMT_2

    # Cap Rate (row 16) - NOI / Purchase Price
    r = 16
    ws.cell(row=r, column=1, value="Cap Rate")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c_idx, col_letter in enumerate(["B", "C", "D"], 2):
        ws.cell(row=r, column=c_idx, value=f"={col_letter}10/{col_letter}4")
        style_body_cell(ws.cell(row=r, column=c_idx), bold=True)
        ws.cell(row=r, column=c_idx).number_format = PCT_FMT_2

    # -----------------------------------------------------------------------
    # 5-Year IRR Section
    # -----------------------------------------------------------------------
    r = 18
    style_section_label(ws.cell(row=r, column=1), "5-Year IRR Estimate")
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(2, 5):
        ws.cell(row=r, column=c).border = THIN_BORDER

    # Exit Cap Rate (row 19)
    r = 19
    ws.cell(row=r, column=1, value="Exit Cap Rate")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c in range(2, 5):
        ws.cell(row=r, column=c, value=0.07)
        style_body_cell(ws.cell(row=r, column=c), input_cell=True)
        ws.cell(row=r, column=c).number_format = PCT_FMT_2

    # Annual NOI Growth (row 20)
    r = 20
    ws.cell(row=r, column=1, value="Annual NOI Growth")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c in range(2, 5):
        ws.cell(row=r, column=c, value=0.03)
        style_body_cell(ws.cell(row=r, column=c), input_cell=True)
        ws.cell(row=r, column=c).number_format = PCT_FMT_2

    # Year 5 NOI (row 21) = NOI * (1 + growth)^5
    r = 21
    ws.cell(row=r, column=1, value="Year 5 NOI")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c_idx, col_letter in enumerate(["B", "C", "D"], 2):
        ws.cell(row=r, column=c_idx,
                value=f"={col_letter}10*(1+{col_letter}20)^5")
        style_body_cell(ws.cell(row=r, column=c_idx))
        ws.cell(row=r, column=c_idx).number_format = CURRENCY_FMT

    # Exit Sale Price (row 22) = Year 5 NOI / Exit Cap Rate
    r = 22
    ws.cell(row=r, column=1, value="Exit Sale Price (Year 5)")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c_idx, col_letter in enumerate(["B", "C", "D"], 2):
        ws.cell(row=r, column=c_idx,
                value=f"={col_letter}21/{col_letter}19")
        style_body_cell(ws.cell(row=r, column=c_idx))
        ws.cell(row=r, column=c_idx).number_format = CURRENCY_FMT

    # Net Proceeds after Loan Payoff (row 23)
    # Approximate remaining loan balance after 5 years using FV
    # =FV(rate/12, 5*12, PMT(rate/12, term*12, -loan), -loan) + exit price
    # Simpler: Exit Price - Remaining Balance
    # Remaining balance = FV(rate/12, 60, -PMT(rate/12,360,loan), -loan)
    r = 23
    ws.cell(row=r, column=1, value="Remaining Loan Balance (Year 5)")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c_idx, col_letter in enumerate(["B", "C", "D"], 2):
        # FV of loan after 60 payments
        formula = f"=FV({col_letter}8/12,60,-PMT({col_letter}8/12,{col_letter}9*12,{col_letter}7),{col_letter}7)"
        ws.cell(row=r, column=c_idx, value=formula)
        style_body_cell(ws.cell(row=r, column=c_idx))
        ws.cell(row=r, column=c_idx).number_format = CURRENCY_FMT

    # Equity at Sale (row 24) = Exit Price - Remaining Balance
    r = 24
    ws.cell(row=r, column=1, value="Equity at Sale")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c_idx, col_letter in enumerate(["B", "C", "D"], 2):
        ws.cell(row=r, column=c_idx,
                value=f"={col_letter}22-{col_letter}23")
        style_body_cell(ws.cell(row=r, column=c_idx))
        ws.cell(row=r, column=c_idx).number_format = CURRENCY_FMT

    # Annual Cash Flow (row 25) = NOI - Debt Service (year 1 approximation)
    r = 25
    ws.cell(row=r, column=1, value="Annual Cash Flow (Year 1)")
    style_body_cell(ws.cell(row=r, column=1), bold=True)
    for c_idx, col_letter in enumerate(["B", "C", "D"], 2):
        ws.cell(row=r, column=c_idx,
                value=f"={col_letter}10-{col_letter}13")
        style_body_cell(ws.cell(row=r, column=c_idx))
        ws.cell(row=r, column=c_idx).number_format = CURRENCY_FMT

    # 5-Year IRR using IRR function (row 26)
    # Cash flows: Year 0 = -Down Payment, Years 1-4 = Cash Flow (growing),
    # Year 5 = Cash Flow + Equity at Sale
    # We build a helper row for the IRR array
    r = 26
    ws.cell(row=r, column=1, value="5-Year IRR (Approx.)")
    style_body_cell(ws.cell(row=r, column=1), bold=True)

    # For IRR, we need to lay out cash flows in helper rows below
    # Rows 28-33: Year 0 through Year 5 cash flows
    # Then row 26 references the IRR of those cells

    # Year labels (row 28)
    r_cf_start = 28
    ws.cell(row=r_cf_start, column=1, value="Year 0 (Investment)")
    ws.cell(row=r_cf_start + 1, column=1, value="Year 1 Cash Flow")
    ws.cell(row=r_cf_start + 2, column=1, value="Year 2 Cash Flow")
    ws.cell(row=r_cf_start + 3, column=1, value="Year 3 Cash Flow")
    ws.cell(row=r_cf_start + 4, column=1, value="Year 4 Cash Flow")
    ws.cell(row=r_cf_start + 5, column=1, value="Year 5 Cash Flow + Sale")

    for label_row in range(r_cf_start, r_cf_start + 6):
        style_body_cell(ws.cell(row=label_row, column=1))

    for c_idx, col_letter in enumerate(["B", "C", "D"], 2):
        # Year 0: -Down Payment
        ws.cell(row=r_cf_start, column=c_idx, value=f"=-{col_letter}6")
        style_body_cell(ws.cell(row=r_cf_start, column=c_idx))
        ws.cell(row=r_cf_start, column=c_idx).number_format = CURRENCY_FMT

        # Year 1: NOI - Debt Service
        ws.cell(row=r_cf_start + 1, column=c_idx,
                value=f"={col_letter}10-{col_letter}13")
        style_body_cell(ws.cell(row=r_cf_start + 1, column=c_idx))
        ws.cell(row=r_cf_start + 1, column=c_idx).number_format = CURRENCY_FMT

        # Year 2: (NOI*(1+growth)^1) - Debt Service
        ws.cell(row=r_cf_start + 2, column=c_idx,
                value=f"={col_letter}10*(1+{col_letter}20)^1-{col_letter}13")
        style_body_cell(ws.cell(row=r_cf_start + 2, column=c_idx))
        ws.cell(row=r_cf_start + 2, column=c_idx).number_format = CURRENCY_FMT

        # Year 3: (NOI*(1+growth)^2) - Debt Service
        ws.cell(row=r_cf_start + 3, column=c_idx,
                value=f"={col_letter}10*(1+{col_letter}20)^2-{col_letter}13")
        style_body_cell(ws.cell(row=r_cf_start + 3, column=c_idx))
        ws.cell(row=r_cf_start + 3, column=c_idx).number_format = CURRENCY_FMT

        # Year 4: (NOI*(1+growth)^3) - Debt Service
        ws.cell(row=r_cf_start + 4, column=c_idx,
                value=f"={col_letter}10*(1+{col_letter}20)^3-{col_letter}13")
        style_body_cell(ws.cell(row=r_cf_start + 4, column=c_idx))
        ws.cell(row=r_cf_start + 4, column=c_idx).number_format = CURRENCY_FMT

        # Year 5: (NOI*(1+growth)^4) - Debt Service + Equity at Sale
        ws.cell(row=r_cf_start + 5, column=c_idx,
                value=f"={col_letter}10*(1+{col_letter}20)^4-{col_letter}13+{col_letter}24")
        style_body_cell(ws.cell(row=r_cf_start + 5, column=c_idx))
        ws.cell(row=r_cf_start + 5, column=c_idx).number_format = CURRENCY_FMT

        # IRR formula in row 26
        ws.cell(row=26, column=c_idx,
                value=f"=IRR({col_letter}{r_cf_start}:{col_letter}{r_cf_start + 5})")
        style_body_cell(ws.cell(row=26, column=c_idx), bold=True)
        ws.cell(row=26, column=c_idx).number_format = PCT_FMT_2

    # Section label for cash flow detail
    r = 27
    style_section_label(ws.cell(row=r, column=1), "IRR Cash Flow Detail")
    ws.cell(row=r, column=1).border = THIN_BORDER
    for c in range(2, 5):
        ws.cell(row=r, column=c).border = THIN_BORDER

    # Freeze top rows
    ws.freeze_panes = "A3"

    return ws


# ===========================================================================
# Main
# ===========================================================================
def main():
    wb = Workbook()

    create_comparable_sales_sheet(wb)
    create_valuation_scenarios_sheet(wb)
    create_buyer_underwriting_sheet(wb)

    filepath = output_path("06_valuation_comps.xlsx")
    wb.save(filepath)
    print(f"Created 06_valuation_comps.xlsx at {filepath}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
