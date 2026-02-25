"""
Generate 11_proforma_3yr.xlsx for Palm Bay Palms Apartments case study.
Three-year pro forma with assumptions section and Excel formula-driven projections.
"""
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data import *

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Reusable style objects
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Arial", size=10, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
BLUE_FONT = Font(name="Arial", size=10, color="1565C0")
BLUE_BOLD_FONT = Font(name="Arial", size=10, bold=True, color="1565C0")
ASSUMPTION_FILL = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
CURRENCY_FMT = '$#,##0'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row, num_cols):
    """Apply navy background, white bold font to a header row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_cell_style(cell, bold=False, blue=False):
    """Apply font and border to a cell."""
    if blue and bold:
        cell.font = BLUE_BOLD_FONT
    elif blue:
        cell.font = BLUE_FONT
    elif bold:
        cell.font = BOLD_FONT
    else:
        cell.font = BODY_FONT
    cell.border = THIN_BORDER


# ===========================================================================
# Build the 3-Year Pro Forma worksheet
# ===========================================================================
def build_proforma(wb):
    ws = wb.active
    ws.title = "3-Year Pro Forma"

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    # -----------------------------------------------------------------------
    # ASSUMPTIONS SECTION  (rows 1-10, light yellow background)
    # -----------------------------------------------------------------------
    # Row 1: title
    ws.cell(row=1, column=1, value="ASSUMPTIONS")
    ws.cell(row=1, column=1).font = Font(name="Arial", size=11, bold=True, color="1565C0")
    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = ASSUMPTION_FILL

    # Assumption input cells  (labels in A, values in B)
    # B2=rent_growth, B3=expense_growth, B4=Y1_vacancy, B5=Y2+_vacancy,
    # B6=Y1_capex, B7=Y2+_capex, B8=mgmt_fee%, B9=debt_service
    assumptions = [
        (2,  "Rent Growth Rate",      0.03,                PCT_FMT),
        (3,  "Expense Growth Rate",   0.02,                PCT_FMT),
        (4,  "Year 1 Vacancy Rate",   0.111,               PCT_FMT),
        (5,  "Year 2+ Vacancy Rate",  0.05,                PCT_FMT),
        (6,  "Year 1 CapEx",          TOTAL_CAPEX,         CURRENCY_FMT),   # 87300
        (7,  "Year 2+ CapEx",         10000,               CURRENCY_FMT),
        (8,  "Management Fee",        MGMT_FEE_PCT,        PCT_FMT),        # 0.08
        (9,  "Annual Debt Service",   ANNUAL_DEBT_SERVICE,  CURRENCY_FMT),  # 73800
    ]

    for row_num, label, value, fmt in assumptions:
        lc = ws.cell(row=row_num, column=1, value=label)
        apply_cell_style(lc, bold=True)
        lc.fill = ASSUMPTION_FILL

        vc = ws.cell(row=row_num, column=2, value=value)
        apply_cell_style(vc, blue=True)
        vc.number_format = fmt
        vc.fill = ASSUMPTION_FILL

        # Fill C & D with assumption background
        for col in range(3, 5):
            c = ws.cell(row=row_num, column=col)
            c.fill = ASSUMPTION_FILL

    # Row 10: blank spacer with assumption fill
    for col in range(1, 5):
        ws.cell(row=10, column=col).fill = ASSUMPTION_FILL

    # -----------------------------------------------------------------------
    # PRO FORMA TABLE  (starts at row 12)
    # -----------------------------------------------------------------------
    HDR_ROW = 12
    for c, text in enumerate(["", "Year 1", "Year 2", "Year 3"], 1):
        ws.cell(row=HDR_ROW, column=c, value=text)
    style_header_row(ws, HDR_ROW, 4)

    # Mutable row counter
    row = [HDR_ROW + 1]   # list so nested functions can mutate

    # --- helper: write one data row ---
    def put(label, y1, y2, y3, fmt=CURRENCY_FMT, bold=False, indent=False):
        r = row[0]
        lbl = ("  " + label) if indent else label
        lc = ws.cell(row=r, column=1, value=lbl)
        apply_cell_style(lc, bold=bold)

        for col, val in [(2, y1), (3, y2), (4, y3)]:
            cell = ws.cell(row=r, column=col, value=val)
            apply_cell_style(cell, bold=bold)
            cell.number_format = fmt

        row[0] += 1
        return r

    # --- helper: section header (bold + underline) ---
    def section(label):
        r = row[0]
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = Font(name="Arial", size=10, bold=True, underline="single")
        cell.border = THIN_BORDER
        for col in range(2, 5):
            ws.cell(row=r, column=col).border = THIN_BORDER
        row[0] += 1
        return r

    # --- helper: blank separator row ---
    def blank():
        r = row[0]
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = THIN_BORDER
        row[0] += 1

    # --- helper: write an expense line that grows at expense_growth ---
    def expense_line(label, base_val):
        """Write an expense item: Y1=base, Y2=Y1*(1+B3), Y3=Y2*(1+B3)."""
        r = put(label, base_val, None, None, indent=True)
        ws.cell(row=r, column=3).value = f"=B{r}*(1+$B$3)"
        ws.cell(row=r, column=4).value = f"=C{r}*(1+$B$3)"
        return r

    # ========================= INCOME ========================= #
    section("INCOME")

    # Gross Potential Rent:  Y1=$298,800  Y2=Y1*(1+rent_growth)  Y3=Y2*(1+rent_growth)
    gpr = put("Gross Potential Rent", GPR_PROFORMA, None, None, indent=True)
    ws.cell(row=gpr, column=3).value = f"=B{gpr}*(1+$B$2)"
    ws.cell(row=gpr, column=4).value = f"=C{gpr}*(1+$B$2)"

    # Less: Vacancy:  Y1=-GPR*B4   Y2=-GPR*B5   Y3=-GPR*B5
    vac = put("Less: Vacancy",
              f"=-B{gpr}*$B$4",
              f"=-C{gpr}*$B$5",
              f"=-D{gpr}*$B$5",
              indent=True)

    # Effective Gross Income = GPR + Vacancy  (vacancy is negative)
    egi = put("Effective Gross Income",
              f"=B{gpr}+B{vac}",
              f"=C{gpr}+C{vac}",
              f"=D{gpr}+D{vac}",
              indent=True)

    # Laundry Income:  $6,000 base, grows at rent_growth
    laun = put("Laundry Income", 6000, None, None, indent=True)
    ws.cell(row=laun, column=3).value = f"=B{laun}*(1+$B$2)"
    ws.cell(row=laun, column=4).value = f"=C{laun}*(1+$B$2)"

    # Late Fees/Other:  $3,000 base, grows at rent_growth
    late = put("Late Fees/Other", 3000, None, None, indent=True)
    ws.cell(row=late, column=3).value = f"=B{late}*(1+$B$2)"
    ws.cell(row=late, column=4).value = f"=C{late}*(1+$B$2)"

    # Total Revenue  (BOLD subtotal)
    trev = put("Total Revenue",
               f"=B{egi}+B{laun}+B{late}",
               f"=C{egi}+C{laun}+C{late}",
               f"=D{egi}+D{laun}+D{late}",
               bold=True)

    blank()

    # ========================= EXPENSES ========================= #
    section("EXPENSES")

    # Each standard expense: base value in Y1, grows at expense_growth (B3)
    tax  = expense_line("Property Taxes",        18750)
    ins  = expense_line("Insurance",             32400)
    rm   = expense_line("Repairs & Maintenance", 22000)

    # Property Management:  =B8 * Total_Revenue  (not expense-growth based)
    mgmt = put("Property Management",
               f"=$B$8*B{trev}",
               f"=$B$8*C{trev}",
               f"=$B$8*D{trev}",
               indent=True)

    util = expense_line("Utilities",              8400)
    land = expense_line("Landscaping",            4800)
    pest = expense_line("Pest Control",           2160)
    lgl  = expense_line("Legal/Admin",            2400)
    rsv  = expense_line("Reserves",               4500)

    # Total Expenses  (SUM of all expense rows, BOLD)
    all_exp = [tax, ins, rm, mgmt, util, land, pest, lgl, rsv]

    def sum_exp(col):
        refs = ",".join(f"{col}{er}" for er in all_exp)
        return f"=SUM({refs})"

    texp = put("Total Expenses",
               sum_exp("B"), sum_exp("C"), sum_exp("D"),
               bold=True)

    blank()

    # ========================= BOTTOM LINE ========================= #
    section("BOTTOM LINE")

    # NOI = Total Revenue - Total Expenses  (BOLD)
    noi = put("Net Operating Income (NOI)",
              f"=B{trev}-B{texp}",
              f"=C{trev}-C{texp}",
              f"=D{trev}-D{texp}",
              bold=True)

    # Less: Capital Expenditures  (negative: Y1=-B6, Y2/3=-B7)
    capex = put("Less: Capital Expenditures",
                f"=-$B$6", f"=-$B$7", f"=-$B$7",
                indent=True)

    # Less: Debt Service  (negative: =-B9 each year)
    ds = put("Less: Debt Service",
             f"=-$B$9", f"=-$B$9", f"=-$B$9",
             indent=True)

    # Cash Flow After Debt Service & CapEx  (BOLD)
    # = NOI + CapEx_row + DS_row   (CapEx and DS are already negative)
    cf = put("Cash Flow After DS & CapEx",
             f"=B{noi}+B{capex}+B{ds}",
             f"=C{noi}+C{capex}+C{ds}",
             f"=D{noi}+D{capex}+D{ds}",
             bold=True)

    # -----------------------------------------------------------------------
    # Freeze panes below the table header
    # -----------------------------------------------------------------------
    ws.freeze_panes = f"A{HDR_ROW + 1}"

    return ws


# ===========================================================================
# Main
# ===========================================================================
def main():
    wb = Workbook()
    build_proforma(wb)

    filepath = output_path("11_proforma_3yr.xlsx")
    wb.save(filepath)
    print(f"Created 11_proforma_3yr.xlsx at {filepath}")


if __name__ == "__main__":
    main()
