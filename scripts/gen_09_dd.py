"""
Generate 09_due_diligence_tracker.xlsx for Palm Bay Palms Apartments case study.
Tracks all due diligence items with status, dates, and conditional formatting.
"""
import sys
import os
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data import *

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Reusable style objects
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Arial", size=10, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
BOLD_RED_FONT = Font(name="Arial", size=10, bold=True, color="FF0000")
DATE_FMT = "YYYY-MM-DD"
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Conditional fill colours for Status column
CLEARED_FILL = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
REVIEWED_FILL = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
RECEIVED_FILL = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
PENDING_FILL = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
REQUESTED_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

STATUS_FILLS = {
    "Cleared": CLEARED_FILL,
    "Reviewed": REVIEWED_FILL,
    "Received": RECEIVED_FILL,
    "Pending": PENDING_FILL,
    "Requested": REQUESTED_FILL,
}


def style_header_row(ws, num_cols):
    """Apply navy background, white bold font to row 1."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_body_cell(cell):
    """Apply default body font and border."""
    cell.font = BODY_FONT
    cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Notes lookup
# ---------------------------------------------------------------------------
NOTES_MAP = {
    "HOA Documents": "N/A - Not within HOA community",
    "Estoppel Letters (16 tenants)": "16 occupied units",
    "Tax Returns (2 years)": "2024 and 2025 returns",
    "Phase I Environmental": "To be ordered by buyer",
    "Trailing 12-Month P&L": "Reviewed - satisfactory",
    "Current Rent Roll": "Reviewed - satisfactory",
    "Insurance Policy (current)": "Reviewed - satisfactory",
    "All Lease Agreements (16)": "Reviewed - all current",
    "Security Deposit Ledger": "Reviewed - matches rent roll",
    "Property Tax Bills (3 years)": "Reviewed - no outstanding liens",
    "Title Commitment": "Reviewed - clear title confirmed",
    "Inspection Report": "Reviewed - see report for details",
    "LLC Operating Agreement": "Reviewed - standard terms",
    "Articles of Organization": "Reviewed - in good standing",
    "Property Management Agreement": "Reviewed - month-to-month",
    "Flood Zone Certification": "Zone X confirmed - minimal risk",
    "Lead Paint Disclosure": "Pre-1978 building - disclosure on file",
    "Utility Bills (12 months)": "Requested from seller",
    "Vendor Contracts": "Awaiting seller response",
    "Survey": "Ordered from title company",
    "Zoning Verification Letter": "Requested from city planning",
    "Certificate of Occupancy": "Awaiting seller response",
    "Building Permits (history)": "Awaiting seller response",
    "Appraisal": "To be ordered by buyer",
    "Bank Statements (6 months)": "Requested from seller",
    "Loan Payoff Letter": "Requested from seller's lender",
    "Roof Warranty": "Requested from seller",
    "HVAC Service Records": "Requested from seller",
    "Fire Inspection Report": "Awaiting seller response",
    "ADA Compliance Documentation": "Awaiting seller response",
}


# ---------------------------------------------------------------------------
# Date generation helpers
# ---------------------------------------------------------------------------
def get_dates_for_item(status, index):
    """Return (date_requested, date_received) based on status."""
    if status in ("Received", "Cleared"):
        date_requested = date(2026, 2, 1)
        # Spread received dates from Feb 5 to Feb 10
        day_offset = 5 + (index % 6)  # cycles through 5-10
        date_received = date(2026, 2, day_offset)
        return date_requested, date_received
    elif status == "Reviewed":
        date_requested = date(2026, 2, 1)
        day_offset = 5 + (index % 6)
        date_received = date(2026, 2, day_offset)
        return date_requested, date_received
    elif status == "Requested":
        date_requested = date(2026, 2, 1)
        return date_requested, None
    else:  # Pending
        return None, None


# ===========================================================================
# Main
# ===========================================================================
def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Due Diligence Tracker"

    # -- Headers ---------------------------------------------------------------
    headers = [
        "DD Item",
        "Category",
        "Status",
        "Responsible Party",
        "Date Requested",
        "Date Received",
        "Days Outstanding",
        "Notes",
        "Critical?",
    ]
    col_widths = [30, 14, 12, 14, 14, 14, 16, 30, 10]

    for c, header in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=header)

    style_header_row(ws, len(headers))

    # -- Data rows -------------------------------------------------------------
    for i, item in enumerate(DD_ITEMS):
        name, category, status, responsible, critical = item
        row = i + 2

        # Column A: DD Item
        ws.cell(row=row, column=1, value=name)

        # Column B: Category
        ws.cell(row=row, column=2, value=category)

        # Column C: Status
        ws.cell(row=row, column=3, value=status)

        # Column D: Responsible Party
        ws.cell(row=row, column=4, value=responsible)

        # Column E & F: Dates
        date_req, date_rec = get_dates_for_item(status, i)

        if date_req is not None:
            cell_e = ws.cell(row=row, column=5, value=date_req)
            cell_e.number_format = DATE_FMT
        else:
            ws.cell(row=row, column=5, value="")

        if date_rec is not None:
            cell_f = ws.cell(row=row, column=6, value=date_rec)
            cell_f.number_format = DATE_FMT
        else:
            ws.cell(row=row, column=6, value="")

        # Column G: Days Outstanding (FORMULA)
        ws.cell(
            row=row, column=7,
            value=f'=IF(F{row}="",TODAY()-E{row},F{row}-E{row})'
        )

        # Column H: Notes
        note = NOTES_MAP.get(name, "")
        ws.cell(row=row, column=8, value=note)

        # Column I: Critical?
        crit_text = "Yes" if critical else "No"
        ws.cell(row=row, column=9, value=crit_text)

        # -- Style all cells in this row ---------------------------------------
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c)
            style_body_cell(cell)

        # -- Conditional fill on Status cell (column C) ------------------------
        status_cell = ws.cell(row=row, column=3)
        if status in STATUS_FILLS:
            status_cell.fill = STATUS_FILLS[status]

        # -- Critical column styling -------------------------------------------
        crit_cell = ws.cell(row=row, column=9)
        if critical:
            crit_cell.font = BOLD_RED_FONT
            crit_cell.border = THIN_BORDER

    # -- Column widths ---------------------------------------------------------
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # -- Freeze top row --------------------------------------------------------
    ws.freeze_panes = "A2"

    # -- Auto-filter -----------------------------------------------------------
    last_row = len(DD_ITEMS) + 1
    ws.auto_filter.ref = f"A1:I{last_row}"

    # -- Save ------------------------------------------------------------------
    filepath = output_path("09_due_diligence_tracker.xlsx")
    wb.save(filepath)
    print(f"Created 09_due_diligence_tracker.xlsx at {filepath}")
    print(f"Total DD items: {len(DD_ITEMS)} data rows")


if __name__ == "__main__":
    main()
