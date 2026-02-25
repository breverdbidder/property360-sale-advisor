"""
Generate 10_closing_worksheet.xlsx for Palm Bay Palms Apartments case study.
Three sheets: Settlement Statement, Security Deposit Transfer, Tenant Notification.
"""
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data import *

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Reusable style objects
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Arial", size=10, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", size=10, bold=True)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color=NAVY)
SUBTITLE_FONT = Font(name="Arial", size=11, bold=True, color=NAVY)
CURRENCY_FMT = '$#,##0'
CURRENCY_FMT_NEG = '$#,##0;($#,##0)'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
BOTTOM_BORDER = Border(bottom=Side(style="thin"))
DOUBLE_BOTTOM = Border(bottom=Side(style="double"))
BOLD_BOTTOM = Border(bottom=Side(style="medium"))

# Occupied units only
OCCUPIED_UNITS = [u for u in UNITS if u[U_STATUS] == "Occupied"]

# Closing assumptions
CLOSING_DATE = "March 15, 2026"
CLOSING_MONTH_DAYS = 31  # March
DAYS_SELLER = 14   # March 1-14
DAYS_BUYER = 17    # March 15-31
ANNUAL_TAXES = PROPERTY["annual_taxes"]  # 18750
DAYS_IN_YEAR = 365
SELLER_TAX_DAYS = 73  # Jan 1 to March 14 = 31 (Jan) + 28 (Feb) + 14 (Mar)
MONTHLY_RENT_COLLECTED = 21000  # approximate


def style_header_row(ws, row, start_col, end_col):
    """Apply navy background, white bold font to a header row."""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_body_cell(cell):
    cell.font = BODY_FONT
    cell.border = THIN_BORDER


def apply_currency(cell):
    cell.number_format = CURRENCY_FMT


def set_cell(ws, row, col, value, font=None, fmt=None, alignment=None, border=None, fill=None):
    """Helper to set a cell value with optional styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    else:
        cell.font = BODY_FONT
    if fmt:
        cell.number_format = fmt
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if fill:
        cell.fill = fill
    return cell


# ===========================================================================
# Sheet 1: Settlement Statement
# ===========================================================================
def create_settlement_statement(wb):
    ws = wb.active
    ws.title = "Settlement Statement"

    # Column widths: A=item code, B=description, C=buyer, D=seller
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    r = 1  # current row tracker

    # ---- Title ----
    set_cell(ws, r, 1, "CLOSING SETTLEMENT STATEMENT", font=TITLE_FONT)
    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    r += 1

    set_cell(ws, r, 1, "Palm Bay Palms Apartments - 2750 Malabar Road SE, Palm Bay, FL 32907",
             font=Font(name="Arial", size=10, italic=True))
    ws.merge_cells(f"A{r}:D{r}")
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    r += 1

    set_cell(ws, r, 1, f"Closing Date: {CLOSING_DATE}",
             font=Font(name="Arial", size=10, italic=True))
    ws.merge_cells(f"A{r}:D{r}")
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    r += 2  # blank row

    # ---- Column headers ----
    header_row = r
    set_cell(ws, r, 1, "", font=HEADER_FONT)
    set_cell(ws, r, 2, "Description", font=HEADER_FONT)
    set_cell(ws, r, 3, "Buyer", font=HEADER_FONT)
    set_cell(ws, r, 4, "Seller", font=HEADER_FONT)
    style_header_row(ws, r, 1, 4)
    r += 1

    # Helper to add a line item row
    def add_line(label, buyer_val=None, seller_val=None, bold=False, indent=0,
                 buyer_formula=None, seller_formula=None):
        nonlocal r
        prefix = "  " * indent
        font = BOLD_FONT if bold else BODY_FONT
        set_cell(ws, r, 1, "", font=font, border=THIN_BORDER)
        set_cell(ws, r, 2, prefix + label, font=font, border=THIN_BORDER)

        c_cell = ws.cell(row=r, column=3)
        if buyer_formula:
            c_cell.value = buyer_formula
        elif buyer_val is not None:
            c_cell.value = buyer_val
        c_cell.font = font
        c_cell.border = THIN_BORDER
        c_cell.number_format = CURRENCY_FMT

        d_cell = ws.cell(row=r, column=4)
        if seller_formula:
            d_cell.value = seller_formula
        elif seller_val is not None:
            d_cell.value = seller_val
        d_cell.font = font
        d_cell.border = THIN_BORDER
        d_cell.number_format = CURRENCY_FMT

        current = r
        r += 1
        return current

    def add_section_header(label):
        nonlocal r
        set_cell(ws, r, 1, "", font=BOLD_FONT, border=THIN_BORDER,
                 fill=PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid"))
        set_cell(ws, r, 2, label, font=BOLD_FONT, border=THIN_BORDER,
                 fill=PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid"))
        set_cell(ws, r, 3, "", font=BOLD_FONT, border=THIN_BORDER,
                 fill=PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid"))
        set_cell(ws, r, 4, "", font=BOLD_FONT, border=THIN_BORDER,
                 fill=PatternFill(start_color="E8EEF4", end_color="E8EEF4", fill_type="solid"))
        r += 1

    def add_blank():
        nonlocal r
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = THIN_BORDER
        r += 1

    # ============================
    # PURCHASE PRICE
    # ============================
    add_section_header("PURCHASE PRICE")
    purchase_row = add_line("Purchase Price", buyer_val=ASKING_PRICE, seller_val=ASKING_PRICE)

    add_blank()

    # ============================
    # BUYER ADJUSTMENTS
    # ============================
    add_section_header("BUYER ADJUSTMENTS")
    earnest_row = add_line("Less: Earnest Money Deposit", buyer_val=-50000, indent=1)

    add_blank()
    add_section_header("PRORATED ITEMS")

    # Property tax proration
    # Seller owes Jan 1 - Mar 14 = 73 days. This is a CREDIT to buyer (seller pays buyer).
    # Formula: annual_taxes * 73 / 365 -- seller's share they already "owe"
    # But from buyer's perspective: buyer gets credited for seller's portion
    # Actually: taxes prorate so that seller credit to buyer = taxes * seller_days / 365
    # Buyer debit = taxes * buyer_days / 365
    # Net: buyer gets credit of seller's portion
    # Let's track debits (+) and credits (-) for buyer
    # Tax credit to buyer: seller pays buyer for Jan1-Mar14
    tax_proration_row = add_line(
        f"Property Tax Proration (Seller: Jan 1 - Mar 14, {SELLER_TAX_DAYS} days)",
        indent=1,
        buyer_formula=f"=-{ANNUAL_TAXES}*{SELLER_TAX_DAYS}/{DAYS_IN_YEAR}",
        seller_formula=f"=-{ANNUAL_TAXES}*{SELLER_TAX_DAYS}/{DAYS_IN_YEAR}",
    )

    # Rent proration - seller owes buyer rent from Mar 15-31 = 17 days
    # Monthly rent ~$21,000; buyer credit = 21000 * 17/31
    rent_proration_row = add_line(
        f"Rent Proration (Buyer: Mar 15-31, {DAYS_BUYER} days of {CLOSING_MONTH_DAYS})",
        indent=1,
        buyer_formula=f"=-{MONTHLY_RENT_COLLECTED}*{DAYS_BUYER}/{CLOSING_MONTH_DAYS}",
        seller_formula=f"=-{MONTHLY_RENT_COLLECTED}*{DAYS_BUYER}/{CLOSING_MONTH_DAYS}",
    )

    # Security deposit transfer - credit to buyer (liability assumed)
    sec_dep_row = add_line(
        "Security Deposits Transferred to Buyer",
        indent=1,
        buyer_formula=f"=-{TOTAL_SECURITY_DEPOSITS}",
        seller_formula=f"=-{TOTAL_SECURITY_DEPOSITS}",
    )

    add_blank()

    # ============================
    # BUYER CLOSING COSTS
    # ============================
    add_section_header("BUYER CLOSING COSTS")

    # New mortgage = ASKING_PRICE * 75% = $1,462,500 (assumed)
    # Loan origination 1% of new mortgage
    loan_orig_row = add_line(
        "Loan Origination Fee (1% of mortgage)",
        indent=1,
        buyer_formula=f"={ASKING_PRICE}*0.75*0.01",
    )
    appraisal_row = add_line("Appraisal Fee", buyer_val=3500, indent=1)
    inspection_row = add_line("Inspection Fee", buyer_val=2800, indent=1)
    title_search_row = add_line("Title Search Fee", buyer_val=1200, indent=1)
    survey_row = add_line("Survey Fee", buyer_val=2500, indent=1)
    buyer_recording_row = add_line("Recording Fees", buyer_val=250, indent=1)
    # Intangible tax: FL 0.2% of mortgage amount
    intangible_row = add_line(
        "Intangible Tax (FL 0.2% of mortgage)",
        indent=1,
        buyer_formula=f"={ASKING_PRICE}*0.75*0.002",
    )

    add_blank()

    # ============================
    # BUYER TOTAL
    # ============================
    # Buyer total due = Purchase price + all buyer debits - all buyer credits
    # Debits: purchase price, closing costs
    # Credits: earnest money, prorated items (taxes, rent, sec deposits)
    buyer_total_row = add_line(
        "BUYER TOTAL DUE AT CLOSING",
        bold=True,
        buyer_formula=(
            f"=C{purchase_row}"     # purchase price
            f"+C{earnest_row}"      # earnest money (negative)
            f"+C{tax_proration_row}"  # tax proration (negative)
            f"+C{rent_proration_row}"  # rent proration (negative)
            f"+C{sec_dep_row}"       # security deposits (negative)
            f"+C{loan_orig_row}"     # loan origination
            f"+C{appraisal_row}"     # appraisal
            f"+C{inspection_row}"    # inspection
            f"+C{title_search_row}"  # title search
            f"+C{survey_row}"        # survey
            f"+C{buyer_recording_row}"  # recording
            f"+C{intangible_row}"    # intangible tax
        ),
    )
    # Bold double border on the total
    for col in range(1, 5):
        ws.cell(row=buyer_total_row, column=col).border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="medium"), bottom=Side(style="double"),
        )
        ws.cell(row=buyer_total_row, column=col).fill = PatternFill(
            start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
        )

    add_blank()
    add_blank()

    # ============================
    # SELLER SIDE
    # ============================
    add_section_header("SELLER ADJUSTMENTS")

    mortgage_row = add_line(
        "Less: Mortgage Payoff",
        indent=1,
        seller_val=-PROPERTY["current_mortgage"],
    )

    # Commission 5%
    commission_row = add_line(
        "Less: Broker Commission (5%)",
        indent=1,
        seller_formula=f"=-{ASKING_PRICE}*0.05",
    )

    title_ins_row = add_line("Less: Title Insurance", seller_val=-4200, indent=1)
    code_enf_row = add_line("Less: Code Enforcement Lien", seller_val=-1850, indent=1)

    # Documentary stamps FL: $0.70 per $100 = 0.007
    doc_stamps_row = add_line(
        "Less: Documentary Stamps (FL $0.70/$100)",
        indent=1,
        seller_formula=f"=-{ASKING_PRICE}*0.007",
    )

    seller_recording_row = add_line("Less: Recording Fees", seller_val=-250, indent=1)

    add_blank()

    # ============================
    # SELLER NET PROCEEDS
    # ============================
    seller_net_row = add_line(
        "SELLER NET PROCEEDS",
        bold=True,
        seller_formula=(
            f"=D{purchase_row}"          # purchase price (positive)
            f"+D{mortgage_row}"          # mortgage payoff (negative)
            f"+D{commission_row}"        # commission (negative)
            f"+D{title_ins_row}"         # title insurance (negative)
            f"+D{code_enf_row}"          # code enforcement (negative)
            f"+D{doc_stamps_row}"        # doc stamps (negative)
            f"+D{seller_recording_row}"  # recording fees (negative)
            f"+D{tax_proration_row}"     # tax proration (negative - seller owes)
            f"+D{rent_proration_row}"    # rent proration (negative - seller owes)
            f"+D{sec_dep_row}"           # security deposits (negative)
        ),
    )
    # Bold double border on the total
    for col in range(1, 5):
        ws.cell(row=seller_net_row, column=col).border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="medium"), bottom=Side(style="double"),
        )
        ws.cell(row=seller_net_row, column=col).fill = PatternFill(
            start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
        )

    # Freeze panes below header
    ws.freeze_panes = f"A{header_row + 1}"

    return ws


# ===========================================================================
# Sheet 2: Security Deposit Transfer
# ===========================================================================
def create_security_deposit_transfer(wb):
    ws = wb.create_sheet("Security Deposit Transfer")

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18

    r = 1

    # Title
    set_cell(ws, r, 1, "SECURITY DEPOSIT TRANSFER SCHEDULE", font=TITLE_FONT)
    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    r += 1

    set_cell(ws, r, 1, "Palm Bay Palms Apartments - 2750 Malabar Road SE, Palm Bay, FL 32907",
             font=Font(name="Arial", size=10, italic=True))
    ws.merge_cells(f"A{r}:D{r}")
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
    r += 2  # blank row

    # Table headers
    header_row = r
    headers = ["Unit", "Tenant", "Monthly Rent", "Security Deposit"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, 1, 4)
    r += 1

    # Data rows - occupied units only
    first_data_row = r
    for u in OCCUPIED_UNITS:
        set_cell(ws, r, 1, u[U_NUM], border=THIN_BORDER)
        set_cell(ws, r, 2, u[U_TENANT], border=THIN_BORDER)
        c3 = set_cell(ws, r, 3, u[U_RENT], border=THIN_BORDER)
        c3.number_format = CURRENCY_FMT
        c4 = set_cell(ws, r, 4, u[U_DEPOSIT], border=THIN_BORDER)
        c4.number_format = CURRENCY_FMT
        r += 1
    last_data_row = r - 1

    # Total row
    total_row = r
    set_cell(ws, r, 1, "", font=BOLD_FONT, border=Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="medium"), bottom=Side(style="double"),
    ))
    set_cell(ws, r, 2, "TOTAL", font=BOLD_FONT, border=Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="medium"), bottom=Side(style="double"),
    ))
    c3 = set_cell(ws, r, 3, f"=SUM(C{first_data_row}:C{last_data_row})", font=BOLD_FONT,
                   border=Border(
                       left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="medium"), bottom=Side(style="double"),
                   ))
    c3.number_format = CURRENCY_FMT
    c4 = set_cell(ws, r, 4, f"=SUM(D{first_data_row}:D{last_data_row})", font=BOLD_FONT,
                   border=Border(
                       left=Side(style="thin"), right=Side(style="thin"),
                       top=Side(style="medium"), bottom=Side(style="double"),
                   ))
    c4.number_format = CURRENCY_FMT
    r += 2  # blank row

    # Acknowledgment text block
    ack_text = (
        "SECURITY DEPOSIT TRANSFER ACKNOWLEDGMENT\n\n"
        f"The undersigned Buyer acknowledges receipt of the above-listed security deposits "
        f"totaling ${TOTAL_SECURITY_DEPOSITS:,.0f}, which Buyer assumes responsibility for "
        f"per Florida Statute 83.49. Buyer agrees to hold deposits in compliance with "
        f"FL 83.49 and to notify all tenants within 30 days of the new deposit holding "
        f"information."
    )
    ack_cell = set_cell(ws, r, 1, ack_text, font=BODY_FONT)
    ws.merge_cells(f"A{r}:D{r + 5}")
    ack_cell.alignment = Alignment(wrap_text=True, vertical="top")
    r += 7

    # Signature lines
    set_cell(ws, r, 1, "Buyer: _________________________", font=BODY_FONT)
    set_cell(ws, r, 3, "Date: _______", font=BODY_FONT)
    r += 2
    set_cell(ws, r, 1, "Seller: _________________________", font=BODY_FONT)
    set_cell(ws, r, 3, "Date: _______", font=BODY_FONT)

    # Freeze panes
    ws.freeze_panes = f"A{header_row + 1}"

    return ws


# ===========================================================================
# Sheet 3: Tenant Notification
# ===========================================================================
def create_tenant_notification(wb):
    ws = wb.create_sheet("Tenant Notification")

    # Column widths
    ws.column_dimensions["A"].width = 80

    r = 1

    # Title
    set_cell(ws, r, 1, "TENANT NOTIFICATION TEMPLATE", font=TITLE_FONT)
    r += 1
    set_cell(ws, r, 1, "Per Florida Statute 83.50 - Change of Ownership Notice",
             font=Font(name="Arial", size=10, italic=True))
    r += 2

    # Template letter
    letter = (
        "[DATE]\n\n"
        "Dear [TENANT NAME],\n\n"
        "RE: Change of Ownership \u2014 Unit [UNIT NUMBER], "
        "2750 Malabar Road SE, Palm Bay, FL 32907\n\n"
        "This letter is to notify you that effective [CLOSING DATE], ownership of the "
        "property at the above address has been transferred from Sunshine Palms Holdings "
        "LLC to [NEW OWNER NAME].\n\n"
        "Pursuant to Florida Statute 83.50, please be advised of the following:\n\n"
        "1. Your new landlord is: [NEW OWNER NAME]\n"
        "2. Your new rent payment address is: [NEW ADDRESS]\n"
        "3. Your security deposit of $[DEPOSIT AMOUNT] has been transferred to and is "
        "now held by the new owner.\n"
        "4. All terms and conditions of your existing lease remain in full force and effect.\n"
        "5. Your next rent payment should be made to the new owner at the address above.\n\n"
        "If you have any questions, please contact [NEW OWNER CONTACT] at [PHONE/EMAIL].\n\n"
        "Sincerely,\n\n"
        "_________________________\n"
        "[NEW OWNER NAME]\n"
        "New Property Owner"
    )

    letter_cell = set_cell(ws, r, 1, letter, font=BODY_FONT)
    # Merge across columns for readability
    ws.merge_cells(f"A{r}:A{r + 24}")
    letter_cell.alignment = Alignment(wrap_text=True, vertical="top")
    r += 26  # move past the letter block

    # Blank row
    r += 1

    # Tenant Reference Table
    set_cell(ws, r, 1, "TENANT REFERENCE TABLE", font=SUBTITLE_FONT)
    r += 1

    # Expand columns for the table
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    # Table headers
    table_header_row = r
    table_headers = ["Unit", "Tenant Name", "Security Deposit", "Lease Expiration"]
    for c, h in enumerate(table_headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, 1, 4)
    r += 1

    # Occupied unit data
    for u in OCCUPIED_UNITS:
        set_cell(ws, r, 1, u[U_NUM], border=THIN_BORDER)
        set_cell(ws, r, 2, u[U_TENANT], border=THIN_BORDER)
        c3 = set_cell(ws, r, 3, u[U_DEPOSIT], border=THIN_BORDER)
        c3.number_format = CURRENCY_FMT
        set_cell(ws, r, 4, u[U_LEASE_END], border=THIN_BORDER)
        r += 1

    # Freeze panes
    ws.freeze_panes = f"A{table_header_row + 1}"

    return ws


# ===========================================================================
# Main
# ===========================================================================
def main():
    wb = Workbook()

    create_settlement_statement(wb)
    create_security_deposit_transfer(wb)
    create_tenant_notification(wb)

    filepath = output_path("10_closing_worksheet.xlsx")
    wb.save(filepath)
    print(f"Created 10_closing_worksheet.xlsx at {filepath}")

    # Quick verification
    from openpyxl import load_workbook
    verify_wb = load_workbook(filepath)
    print(f"Sheets: {verify_wb.sheetnames}")
    for name in verify_wb.sheetnames:
        ws = verify_wb[name]
        print(f"  '{name}': {ws.max_row} rows x {ws.max_column} columns")


if __name__ == "__main__":
    main()
