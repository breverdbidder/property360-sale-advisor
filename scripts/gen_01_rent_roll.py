"""
Generate 01_rent_roll_2025.xlsx for Palm Bay Palms Apartments case study.
Three sheets: Rent Roll, Monthly Collections, Summary.
"""
import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data import *

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, numbers, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Reusable style objects
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Arial", size=10, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
VACANT_FILL = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
BELOW_MARKET_FILL = PatternFill(start_color=YELLOW_BG, end_color=YELLOW_BG, fill_type="solid")
CURRENCY_FMT = '$#,##0'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, num_cols):
    """Apply navy background, white bold font to row 1."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_body_cell(cell):
    """Apply default body font and border."""
    cell.font = BODY_FONT
    cell.border = THIN_BORDER


def apply_currency(cell):
    """Apply currency number format to a cell."""
    cell.number_format = CURRENCY_FMT


# ===========================================================================
# Sheet 1: Rent Roll
# ===========================================================================
def create_rent_roll_sheet(wb):
    ws = wb.active
    ws.title = "Rent Roll"

    headers = [
        "Unit #", "Unit Type", "SF", "Tenant Name", "Lease Start",
        "Lease Expiration", "Monthly Rent", "Market Rent", "Rent Delta",
        "Security Deposit", "Status", "Delinquent?", "Notes",
    ]
    col_widths = [8, 10, 6, 25, 12, 14, 12, 12, 12, 14, 10, 16, 20]

    # Write headers
    for c, header in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=header)

    style_header_row(ws, len(headers))

    # Write unit data (rows 2-19)
    for i, u in enumerate(UNITS):
        row = i + 2
        ws.cell(row=row, column=1, value=u[U_NUM])
        ws.cell(row=row, column=2, value=u[U_TYPE])
        ws.cell(row=row, column=3, value=u[U_SF])
        ws.cell(row=row, column=4, value=u[U_TENANT])
        ws.cell(row=row, column=5, value=u[U_LEASE_START])
        ws.cell(row=row, column=6, value=u[U_LEASE_END])
        ws.cell(row=row, column=7, value=u[U_RENT])
        ws.cell(row=row, column=8, value=u[U_MARKET])
        # Column I: Rent Delta as FORMULA
        ws.cell(row=row, column=9, value=f"=H{row}-G{row}")
        ws.cell(row=row, column=10, value=u[U_DEPOSIT])
        ws.cell(row=row, column=11, value=u[U_STATUS])
        ws.cell(row=row, column=12, value=u[U_DELINQ])
        ws.cell(row=row, column=13, value=u[U_NOTES])

        # Style every body cell in the row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=c)
            style_body_cell(cell)

        # Currency formatting for columns G, H, I, J
        for c in [7, 8, 9, 10]:
            apply_currency(ws.cell(row=row, column=c))

        # Conditional row highlighting
        is_vacant = u[U_STATUS] == "Vacant"
        is_below_market = u[U_RENT] < u[U_MARKET] and not is_vacant

        if is_vacant:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = VACANT_FILL
        elif is_below_market:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = BELOW_MARKET_FILL

    # Column widths
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Freeze pane & auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{len(UNITS) + 1}"

    return ws


# ===========================================================================
# Sheet 2: Monthly Collections
# ===========================================================================
def create_monthly_collections_sheet(wb):
    ws = wb.create_sheet("Monthly Collections")

    # Header row
    headers = ["Unit #"] + MONTHS
    for c, header in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=header)

    style_header_row(ws, len(headers))

    collections = get_monthly_collections()

    # Data rows (rows 2-19)
    for i, u in enumerate(UNITS):
        row = i + 2
        unit_num = u[U_NUM]
        ws.cell(row=row, column=1, value=unit_num)
        style_body_cell(ws.cell(row=row, column=1))

        month_data = collections.get(unit_num, [0] * 12)
        for m in range(12):
            cell = ws.cell(row=row, column=m + 2, value=month_data[m])
            style_body_cell(cell)
            apply_currency(cell)

        # Highlight vacant unit rows
        if u[U_STATUS] == "Vacant":
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).fill = VACANT_FILL

    last_data_row = len(UNITS) + 1  # row 19
    summary_start = last_data_row + 2  # row 21, leave a blank row

    # -- Summary Row: Total Collected --
    total_row = summary_start
    ws.cell(row=total_row, column=1, value="Total Collected")
    style_body_cell(ws.cell(row=total_row, column=1))
    ws.cell(row=total_row, column=1).font = Font(name="Arial", size=10, bold=True)

    for m in range(12):
        col_letter = get_column_letter(m + 2)
        cell = ws.cell(
            row=total_row, column=m + 2,
            value=f"=SUM({col_letter}2:{col_letter}{last_data_row})"
        )
        style_body_cell(cell)
        apply_currency(cell)
        cell.font = Font(name="Arial", size=10, bold=True)

    # -- Summary Row: Vacancy Loss --
    vacancy_row = summary_start + 1
    ws.cell(row=vacancy_row, column=1, value="Vacancy Loss")
    style_body_cell(ws.cell(row=vacancy_row, column=1))
    ws.cell(row=vacancy_row, column=1).font = Font(name="Arial", size=10, bold=True)

    # GPR monthly = sum of all market rents
    gpr_monthly = GROSS_POTENTIAL_RENT_MONTHLY
    for m in range(12):
        col_letter = get_column_letter(m + 2)
        # Vacancy loss = GPR monthly - total collected for that month
        cell = ws.cell(
            row=vacancy_row, column=m + 2,
            value=f"={gpr_monthly}-{col_letter}{total_row}"
        )
        style_body_cell(cell)
        apply_currency(cell)
        cell.font = Font(name="Arial", size=10, bold=True)

    # -- Summary Row: Delinquency Loss --
    delinq_row = summary_start + 2
    ws.cell(row=delinq_row, column=1, value="Delinquency Loss")
    style_body_cell(ws.cell(row=delinq_row, column=1))
    ws.cell(row=delinq_row, column=1).font = Font(name="Arial", size=10, bold=True)

    # Calculate delinquency losses per month from data
    # Unit 103 (row index 2 = row 4): rent $950, missed months 10,11 (Jan 2026, Feb 2026)
    # Unit 207 (row index 12 = row 14): rent $1200, missed month 11 (Feb 2026)
    delinq_units = {}
    for i, u in enumerate(UNITS):
        if u[U_DELINQ] != "No" and u[U_DELINQ] != "":
            delinq_units[u[U_NUM]] = {
                "row": i + 2,
                "rent": u[U_RENT],
                "delinq": u[U_DELINQ],
            }

    # Build per-month delinquency loss from the collections data
    delinq_monthly = [0] * 12
    for unit_num, info in delinq_units.items():
        month_data = collections.get(unit_num, [0] * 12)
        for m in range(12):
            if month_data[m] == 0:
                delinq_monthly[m] += info["rent"]

    for m in range(12):
        cell = ws.cell(row=delinq_row, column=m + 2, value=delinq_monthly[m])
        style_body_cell(cell)
        apply_currency(cell)
        cell.font = Font(name="Arial", size=10, bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 8
    for m in range(12):
        ws.column_dimensions[get_column_letter(m + 2)].width = 12

    # Freeze pane & auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_data_row}"

    return ws


# ===========================================================================
# Sheet 3: Summary
# ===========================================================================
def create_summary_sheet(wb):
    ws = wb.create_sheet("Summary")

    headers = ["Metric", "Value"]
    for c, header in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=header)

    style_header_row(ws, len(headers))

    # Define summary rows: (label, formula, format)
    # Reference range for occupied units in Rent Roll: K2:K19 (Status), G2:G19 (Rent), etc.
    metrics = [
        (
            "Occupancy Rate",
            "=COUNTIF('Rent Roll'!K2:K19,\"Occupied\")/18",
            PCT_FMT,
        ),
        (
            "Average Rent per Unit (Occupied)",
            "=AVERAGEIF('Rent Roll'!K2:K19,\"Occupied\",'Rent Roll'!G2:G19)",
            CURRENCY_FMT,
        ),
        (
            "Average Rent per SF (Occupied)",
            "=SUMPRODUCT(('Rent Roll'!K2:K19=\"Occupied\")*'Rent Roll'!G2:G19)/SUMPRODUCT(('Rent Roll'!K2:K19=\"Occupied\")*'Rent Roll'!C2:C19)",
            '$#,##0.00',
        ),
        (
            "Total Monthly Income",
            "=SUM('Rent Roll'!G2:G19)",
            CURRENCY_FMT,
        ),
        (
            "Total Annual Income",
            "=SUM('Rent Roll'!G2:G19)*12",
            CURRENCY_FMT,
        ),
        (
            "Below-Market Units",
            "=COUNTIF('Rent Roll'!I2:I19,\">\"&0)",
            '0',
        ),
        (
            "Total Monthly Upside (to Market)",
            "=SUM('Rent Roll'!I2:I19)",
            CURRENCY_FMT,
        ),
    ]

    for i, (label, formula, fmt) in enumerate(metrics):
        row = i + 2
        label_cell = ws.cell(row=row, column=1, value=label)
        style_body_cell(label_cell)
        label_cell.font = Font(name="Arial", size=10, bold=True)

        val_cell = ws.cell(row=row, column=2, value=formula)
        style_body_cell(val_cell)
        val_cell.number_format = fmt

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

    # Freeze pane
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:B1"

    return ws


# ===========================================================================
# Main
# ===========================================================================
def main():
    wb = Workbook()

    create_rent_roll_sheet(wb)
    create_monthly_collections_sheet(wb)
    create_summary_sheet(wb)

    filepath = output_path("01_rent_roll_2025.xlsx")
    wb.save(filepath)
    print(f"Created 01_rent_roll_2025.xlsx at {filepath}")


if __name__ == "__main__":
    main()
